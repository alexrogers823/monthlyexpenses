from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl.styles import Font, Style, Color, PatternFill
from openpyxl.cell import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.chart.layout import Layout , ManualLayout
from openpyxl.chart.label import DataLabel, DataLabelList
import openpyxl as px
#import save_and_update as saving
import sys, os, subprocess, time, datetime, random, webbrowser

#Part I: creating/updating excel spreadsheet
# S = sheet (so sumS and JanS are summary sheet and January sheet)
def reviewMaterials():
    print('Login into your accounts')
    if day_num <= 12:
        WF = 23 - day_num
        Dis = 12 - day_num
    elif day_num > 12 and day_num <= 23:
        WF = 23 - day_num
        Dis = '20+'
    else:
        WF = '20+'
        Dis = '10+'
    print('REMINDER: You have %s days to pay off your WF balance,\n and %s days to pay off your Discover balance' % (WF, Dis))
    print()
    #Discover is on the 12th, and WF is on the 23rd
    webbrowser.open('https://www.discover.com')
    webbrowser.open('https://www.wellsfargo.com')
    webbrowser.open('https://www.capitalone.com')
    print('Do you need to see the excel sheet as well? [Y or N]')
    if max_row > 2:
        print('(The last two inputs were %s for $%s, and %s for $%s)' % (month['A'+str(max_row)].value, month['D'+str(max_row)].value, month['A'+str(max_row-1].value, month['D'+str(max_row-1].value))
    if input().lower().startswith("y"):
        opener = 'open'
        print('Reviewing...')
        time.sleep(2)
        subprocess.call([opener, (year_title + '.xlsx')])

#Sets new year automatically
def set_new_year():
    new_year = px.load_workbook('blank.xlsx') #Dont need this anymore bc of try/else below
    new_year.create_sheet(index = 0, title = 'Summary')
    new_year.get_sheet_by_name('Summary')['J1'] = 'Average'
    new_year.create_sheet(index = 1, title = 'Category Words')
    setCategorySheet(new_year.get_sheet_by_name('Category Words'))
    new_year.save(str(int(year)+1) +' Monthly Expenses.xlsx')
    return

def setWorkbook():
    if Sum[get_column_letter(month_num+9) + '1'].value == 'Average':
        if month_num == 12:
            set_new_year()
        wb.create_sheet(index = month_num, title = month_list[month_num])
        shift.Shift(Sum, Sum.max_column, Sum.max_row, month_num+12).right()
    current_month = wb.get_sheet_by_name(month_list[month_num])
    return current_month

def setSummarySheet():
#    yearTotal = Font(name='Calibri', bold=True, color='009999')
    year_style = Style(font=Font(name='Calibri', bold=True, color='009999'))
#    mFont = Font(name='Calibri', color='FFFFFF')
    mStyle = Style(font=Font(name='Calibri', color='FFFFFF'))
    month_fill = PatternFill(start_color='2E2EB8', fill_type='solid')
    category_list = ['Rent', 'Utilities', 'Phone', 'Apparel','Supplies', 'Technology', 'Services', 'Health/Gym', 'Haircut', 'Groceries', 'Other Food', 'Gas & Parking', 'Insurance','Bank & Credit Card', 'Student Loans', 'Vehicle Payments', 'Entertainment', 'Subscriptions', 'Travel', 'Public Transportation', 'Special/Seasonal', 'Other', '', 'Total Expenses']
    for i in range(0, month_num+1):
        Sum[get_column_letter(i+9) + '1'] = month_list[i]
        Sum[get_column_letter(i+9) + '1'].style = month_style

    Sum[get_column_letter(month_num+10) + '1'] = 'Average'
    Sum[get_column_letter(month_num+11) + '1'] = '3-Mo. Avg'
    Sum[get_column_letter(month_num+12) + '1'] = 'Month Goal'
    for a in range(10, 12):
        Sum[get_column_letter(month_num+a) + '1'].style = Style(font=Font(name='Georgia', bold=True, color='AB800A'))

    Sum[get_column_letter(month_num+12)+'1'].style = Style(font=Font(name='Impact', bold=True, color='2E2EB8'))
    Sum[get_column_letter(month_num + 9) +'1'].style = Style(font=Font(name='Georgia', bold=True, color='2E2EB8'))
    for j in range(len(category_list)):
        Sum['H' + str(j+2)] = category_list[j] + ':'
        Sum['H' + str(j+2)].style = month_style
        Sum['I' + str(j+2)].style = year_style
        Sum[get_column_letter(month_num+10) + str(j+2)].style = Style(font=Font(name='Calibri', bold=True, color='AB800A'))
        Sum[get_column_letter(month_num+11) + str(j+2)].style = Style(font=Font(name='Calibri', bold=True, color='AB800A'))
        Sum[get_column_letter(month_num + 9) + str(j+2)].style = mStyle
        Sum[get_column_letter(month_num + 9) + str(j+2)].fill = month_fill
        Sum[get_column_letter(month_num + 9) + str(len(category_list))].fill = PatternFill(fill_type=None)

    for e in range(len(month_list)):
        Sum[get_column_letter(e+8) +str(len(category_list))] = ''
    setCategorySheet(category_words)
    return category_list

def setCategorySheet(tab):
    tab['A1'] = 'Keyword'
    tab['A1'].style = Style(font=Font(name='Georgia', bold=True))
    tab['B1'] = 'Category'
    tab['B1'].style = Style(font=Font(name='Georgia', bold=True))
    if tab != category_words:
        for i in range(2, category_words.max_row + 1):
            tab['A'+str(i)] = category_words['A'+str(i)].value
            tab['B'+str(i)] = category_words['B'+str(i)].value
    return

def setSubCategories():
    category_dict = {}
    for i in range(len(categories)-2):
        category_dict[categories[i]] = []
        for j in range(2, category_words.max_row + 1):
            if category_words['B'+str(j)].value == categories[i]:
                category_dict[categories[i]].append((category_words['A'+str(j)].value).title())
    return category_dict


def formatSheet(month):
    month.column_dimensions['A'].width = 23
    month.column_dimensions['B'].width = 15
    month.column_dimensions['C'].width = 23
    month.column_dimensions['E'].width = 20
    Sum.column_dimensions['H'].width = 23

    if max_row > 29:
        month.freeze_panes = 'A2'

 #   fontmonth_style = Font(name='Georgia', bold=True)
    month_style = Style(font=Font(name='Georgia', bold=True))

    return month_style

def setMonth():
    month['A1'] = 'Expense'
    month['A1'].style = month_style
    month['B1'] = 'Category'
    month['B1'].style = month_style
    month['C1'] = 'Notes'
    month['C1'].style = month_style
    month['D1'] = 'Price'
    month['D1'].style = month_style
    month['E1'] = 'Cumulative Price'
    month['E1'].style = month_style
    month['G1'] = '*Most recent expenses in bold text (Scroll Down)'
    return

def revertStyle():
    standard = Style(font=Font(name='Calibri', bold=False))
    for i in range(2,(max_row+1)):
        for j in range(0,5):
            month[get_column_letter(j+1) + str(i)].style = standard

    for a in range(len(categories)):
        for b in range(1, month_num):
            Sum[get_column_letter(b+9) + str(a+2)].style = standard

def checkCate():
    value = ''
    for i in range(2, category_words.max_row+1):
        if category_words['A'+str(i)].value == keyword:
            value = category_words['B'+str(i)].value

    return value


def addToCate(): #adds new keyword to category upon request
    category_max = category_words.max_row
    category_max += 1
    category_words['A'+str(d)] = keyword
    print('Which category to assign?')
    print()
    for i in range(len(categories)-3):
        print(i, categories[i])
    assign = int(input())
    category_words['B'+str(category_max)] = categories[assign]


def addExpense():
    global max_row, keyword, items
    max_row += 1
    fontCur = Font(name='Cooper Black', bold=False)
    current = Style(font = fontCur)
    Ex1 = random.randint(0, 4)
    Ex2 = random.randint(5, 9)
    Ex3 = random.randint(10, len(categories) - 3)
    Name1 = random.randint(1, max_row-1)
    Name2 = random.randint(1, max_row-1)
    Name3 = random.randint(1, max_row-1)

    if max_row < 5:
        print('What\'s the name of your expense? (Ex: eggs, clothes, vodka)')
    else:
        print('What\'s the name of your expense? (Ex: {}, {}, {})'.format(month['A'+str(Name1)].value, month['A'+str(Name2)].value, month['A'+str(Name3)].value)
    expense = input()
    if expense.lower().startswith('correct'):
        print('Correction initiated')
        shortExit()
        sys.exit()
    if expense.lower() == 'no' or expense.lower() == 'done':
        shortExit()
        sys.exit()
    if expense.lower().endswith('cashback'):
        expense.split()
        expense = expense[0]
        print('How much cash back?')
        autoAdd('cashback', 'Bank', float(input()))
    items += 1
    month['A' + str(max_row)] = expense
    keyword = expense
    isKeyword = checkCate()
    print()
    if isKeyword != '':
        month['B'+str(max_row)] = isKeyword.title()
    else:
        print('Under which category? (Ex: {}, {}, {})'.format(categories[Ex1], categories[Ex2], categories[Ex3]))
        print('(* at end to assign)')
        category = input()
        if category.endswith('*'):
            addToCate()
        category = category.strip('*')
        print()
        if category == 'food':
            if input('Was it food from the grocery store? [Y or N]\n').lower.startswith("y"):
                month['B' +str(max_row)] = 'Groceries'
            else:
                month['B' +str(max_row)] = 'Res'
        elif category == 'cvs':
            if input('Was it health related? (if for supplies then N) [Y or N]\n').lower().startswith("y"):
                month['B' +str(max_row)] = 'Health'
            else:
                month['B' +str(max_row)] = 'Supplies'
        else:
            month['B' + str(max_row)] = category.title()
    print('Additional notes?')
    if day_num < 10:
        print('(if from previous month, put \'qq\' here)')
    month['C' + str(max_row)] = input()
    print()
    while True:
        cost = input('Total cost? (Enter as x.xx)')
        try:
            month['D'+str(max_row)] = float(cost)
        except ValueError:
            print('Be careful! You have to enter a number here')
        else:
            break
    if ['B'+str(max_row)] == 'Other Food' and float(cost) > float(25):
        pay = input('Did you pay for others\' food?')
        if pay.lower().startswith('y'):
            month['D'+str(max_row)].fill = PatternFill(start_color='6DC066', fill_type='solid')
    print()

    for i in range(0,5):
        month[get_column_letter(i+1) + str(max_row)].style = current


def autoAdd(autoname, autocate, cash):
    global max_row, items
    month['A'+str(max_row)] = autoname
    month['B'+str(max_row)] = autocate
    month['D'+str(max_row)] = cash
    for i in range(0,5):
        month[get_column_letter(i+1) + str(max_row)].style = Style(font=Font('Cooper Black', bold=False))
    cumulativePrice()
    max_row += 1
    items += 1
    return



def categoryTotal():
    category_length = len(categories)
    for g in range(1, month_num+1):
        selected_month = wb.get_sheet_by_name(month_list[g])
        column = g+1
        selected_bottom_row = selected_month.max_row
        for i in range(category_length-2):
            Sum[get_column_letter(column+8) + str(i+2)] = 0
            Sum['I'+str(i+2)] = '=SUM(J{0}:{1}{0})'.format(i+2, get_column_letter(month_num+9))

            # Will change to dictionary
            y = i+1
            if y == 1: #Rent
                ct = ['Rent','Lease',]
            if y == 2: #Utilities
                ct = ['Util','Utility','Utilities','Water Bill','Electricity', 'Water', 'Bill']
            if y == 3: #Phone
                ct = ['Phone Bill', 'Phone']
            if y == 4: #Apparel
                ct = ['Apparel', 'Clothes', 'Shoes']
            if y == 5: #Supplies
                ct = ['Supplies', 'Office', 'Bathroom']
            if y == 6: #Tech
                ct = ['Technology', 'Tech'] #...
            if y == 7: #Service
                ct = ['Service', 'Services', 'Dry Cleaning']
            if y == 8: #Health/Gym
                ct = ['Health','Gym','Medical','Med','Gnc','Vitamins', 'Health/Gym']
            if y == 9: #Haircut
                ct = ['Haircut', 'Hair', 'Fade', 'Cut', 'Lineup']
            if y == 10: #Groceries
                ct = ['Groceries','Grocery']
            if y == 11: #Other Food
                ct = ['Other Food', 'Res','Snacks','Candy','Beer','Wine','Alcohol','Liquor','Coffee','Drink']
            if y == 12: #Gas and Parking
                ct = ['Gas', 'Parking', 'Gas & Parking']
            if y == 13: #Insurance
                ct = ['Insurance', 'Ins']
            if y == 14: #Credit Card
                ct = ['Bank & Credit Card', 'Credit','Credit Card','Bank', 'Interest', 'Cc']
            if y == 15: #Student Loans
                ct = ['Student','Loans','Student Loans','Student Loan', 'FAFSA']
            if y == 16: #Vehicle Expense
                ct = ['Car','Motorcycle','Engine']
            if y == 17: #Entertainment
                ct = ['Entertainment','Movies','Sport','Sports','Game']
            if y == 18: #Subscriptions
                ct = ['Sub','Subscriptions','Xxx','Recurring','Spotify']
            if y == 19: #Travel
                ct = ['Travel','Flight','Train','Plane','Airplane','Bus']
            if y == 20: #Public Transportation
                ct = ['Public Transportation', 'Transportation', 'Uber', 'Marta', 'Cab']
            if y == 21: #Seasonal/Special
                ct = ['Special', 'Seasonal', 'Graduation', 'Spring Break', 'Recital', 'Gift', 'Gifts']


            for j in range(2,(selected_bottom_row+1)): #Updates all active categories #category_dict[categories[i]] instead of ct
                if selected_month['B' + str(j)].value in ct and selected_month['C' + str(j)].value != 'qq':
                    add = float(Sum[get_column_letter(column+8) + str(i+2)].value) + float(selected_month['D' + str(j)].value)
                    Sum[get_column_letter(column+8) + str(i+2)] = float(add)
                if g != 1 and selected_month['B' + str(j)].value in ct and selected_month['C' + str(j)].value == 'qq':
                    add = float(Sum[get_column_letter(column+7) + str(i+2)].value) + float(selected_month['D' + str(j)].value)
                    Sum[get_column_letter(column+7) + str(i+2)] = float(add)

            if month_num > 3:
                Sum[get_column_letter(month_num+10) + str(i+2)] = '=ROUND(AVERAGE(J{0}:{1}{0}), 2)'.format(i+2, get_column_letter(month_num+8))
                Sum[get_column_letter(month_num+11) + str(i+2)] = '=ROUND(AVERAGE({1}{0}:{2}{0}), 2)'.format(i+2, get_column_letter(month_num+6), get_column_letter(month_num+8))
            else:
                Sum[get_column_letter(month_num+10) + str(i+2)] = '=\'[{} Monthly Expenses.xlsx]Summary\'!$V${}'.format(int(year)-1, i+2)
                Sum[get_column_letter(month_num+11) + str(i+2)] = '=\'[{} Monthly Expenses.xlsx]Summary\'!$W${}'.format(int(year)-1, i+2)

        #Below updates the 'other' category and total for month
        Sum[get_column_letter(column+8) + str(category_length-1)] = '={0}!E{1}-SUM(Summary!{2}2:{2}{3})'.format(month_list[g], selected_bottom_row, get_column_letter(column+8), category_length-2)
        Sum[get_column_letter(column+8) + str(category_length+1)] = '=SUM({0}2:{0}{1})'.format(get_column_letter(column+8), category_length-1)

    # Updating the year total and average
    Sum['I' +str(category_length-1)] = '=SUM(J{0}:{1}{0})'.format(category_length-1, get_column_letter(month_num+9))
    if month_num > 3:
        Sum[get_column_letter(month_num+10) + str(category_length+1)] = '=ROUND(AVERAGE(J{0}:{1}{0}), 2)'.format(category_length-1, get_column_letter(month_num+8))
        Sum[get_column_letter(month_num+11) + str(category_length+1)] = '=ROUND(AVERAGE({1}{0}:{2}{0}), 2)'.format(category_length+1, get_column_letter(month_num+6), get_column_letter(month_num+8))
    else:
        Sum[get_column_letter(month_num+10) + str(category_length+1)] = '=\'/Users/alexrogers823/Documents/Python Projects/[{} Monthly Expenses.xlsx]Summary\'!$V${}'.format(int(year)-1, category_length+1)
        Sum[get_column_letter(month_num+11) + str(category_length+1)] = '=\'/Users/alexrogers823/Documents/Python Projects/[{} Monthly Expenses.xlsx]Summary\'!$W${}'.format(int(year)-1, category_length+1)

    for r in range(len(categories)):
        if Sum[get_column_letter(month_num+12)+str(r+2)].value != None and Sum[get_column_letter(month_num+9)+str(r+2)].value > Sum[get_column_letter(month_num+12)+str(r+2)].value:
            Sum[get_column_letter(month_num+12)+str(r+2)].style = Style(font=Font(name='Calibri', color='FF0000'))


def cumulativePrice():
    price_cell = max_row - 1
    if month['C'+str(max_row)].value == 'qq':
        if max_row == 2:
            month['E2'] = 'Not Included'
            return
        if max_row > 2:
            month['E'+str(max_row)] = '=E'+str(max_row-1)
            return
    else:
        if max_row == 2:
            month['E2'] = '=D2'
            return
        elif max_row > 2 and month['E'+str(max_row-1)].value == 'Not Included':
            match = False
            while match == False:
                if month['E'+str(price_cell)].value == 'Not Included':
                    if price_cell == 1:
                        month['E'+str(max_row)] = '=D'+str(max_row)
                    else:
                        price_cell -= 1
                else:
                    if month['E'+str(price_cell)].value == 'Cumulative Price':
                        month['E'+str(max_row)] = '=D'+str(max_row)
                    else:
                        month['E'+str(max_row)] = '=E'+str(price_cell)
                        match == True
                    break
            return
        else:
#            month['E'+str(max_row)] = '=D' + str(max_row) + '+E' + str((x) - 1)
            month['E'+str(max_row)] = '=IFERROR(D{0}+E{1},D{0})'.format(str(max_row), str(max_row-1))
            return


def monthChartBreakdown():
    current_month = ['','January','February','March','April','May','June','July','August','September','October','November','December']
    for i in range(1, month_num+1):
        pie = PieChart()
        cat_length = len(categories)
        month_data = Reference(Sum, min_col = i+9, min_row = 2, max_row = cat_length-1)
        labels = Reference(Sum, min_col=8, min_row=2, max_row = cat_length-1)
        pie.add_data(month_data)
        pie.set_categories(labels)
        pie.title = current_month[i]+ ' Expenses by Category'
        pie.width = 18.0
        pie.height = 12.0
        pie.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=0.99, w=0.25))

        wb.get_sheet_by_name(month_list[i]).add_chart(pie, 'G3')

def chartBreakdown():
    pie = PieChart()
    cat_length = len(categories)

    data = Reference(Sum, min_col=9, min_row=2, max_row= cat_length-1)
    labels = Reference(Sum, min_col=8, min_row=2, max_row= cat_length-1)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = 'Breakdown of Expenses'
    pie.width = 15.0
    pie.height = 12.0
    pie.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=0.99, w=0.25))

    Sum.add_chart(pie, 'A1')
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    bar = BarChart()
    barData1 = Reference(Sum, min_col=month_num+9, min_row=1, max_row=cat_length-1)
    barData2 = Reference(Sum, min_col=month_num+12, min_row=1, max_row=cat_length-1)
    bar.add_data(barData1, titles_from_data=True)
    bar.add_data(barData2, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'Goal Comparison'
    bar.width = 2.0*cat_length
    bar.height = 12.0
    bar.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=1.99, w=0.25))
    Sum.add_chart(bar, 'A28')

#def save_and_update():
#    wbSave = saving.Update(xl, 'open')
#    wbSave.save()
#    wbSave.openXL()

def save_and_update():
    wb.save(year_title + '.xlsx')
    print('Would you like to see the additions to your expenses? [Y or N]')
    openXL = input()
    if openXL.lower().startswith('y'):
        opener = 'open'
        print('Opening...')
        time.sleep(2)
        subprocess.call([opener, (year_title + '.xlsx')])

def add_to_database(expense, category, notes, cost):
    data = {}
    data.expense = expense
    data.category = category
    data.other = notes
    data.cost = cost

def secret():
    #Updates utilities and groceries in Monthly Breakdown
    groceries = '=\'[{}]Summary\'$V$3'.format(excel)
    utilities = '=\'[{}]Summary\'$V$11'.format(excel)
    secret_file = px.load_workbook('Life Management.xlsx')
    management = secret_file.get_sheet_by_name('Money Allocation')
    management['B5'] = utilities #utility 3-mo avg
    management['B6'] = groceries #groceries 3-mo avg
    secret_file.save('Life Management.xlsx')


def orderOfStatements():
    while True:
        addExpense()
        cumulativePrice()
        categoryTotal()

def manual_correction(month):
    cell = input("Which cell to change?\n")
    correction = input("Enter correction here: ")
    month[cell].value = correction
    return

def presetGoals():
    goals_arr = []
    for i in range(len(categories)-3):
        if Sum[get_column_letter(month_num+12)+str(i+2)].value == None:
            goals_arr.append('0')
        else:
            goals_arr.append(Sum[get_column_letter(month_num+12)+str(i+2)].value)

    return g


def goals():
    global goals_list
    print('Enter goal for category, followed by number (Haircut, 30)')
    print()
    for i in range(len(goals_list)):
        print('%s , Current goal: %s' % (categories[i], goals_list[i]))
    print()
    while True:
        goal = input()
        if goal.lower().startswith('n'):
            break
        else:
            goal = goal.title().split(', ')
            for c in range(len(goals_list)):
                if categories[c].startswith(goal[0].title()):
                    Sum[get_column_letter(month_num+12)+str(c+2)] = float(goal[1])
                    goals_list[i] = goal[1]
            print('Next one [n to finish]')
 #       if input().lower().startswith('n'):
#            k = False
#            break
#        else:
#            continue

    print('continue with Monthly Expenses?')
    return input().lower()

def overspent_goal():
    overspent = PatternFill(start_color='FF0000', fill_type='solid')
    for i in range(len(categories)-3):
        cell = Sum[get_column_letter(month_num+12)+str(i+2)].value
        if type(cell) == int and Sum[get_column_letter(month_num+9)+str(i+2)].value > cell:
            Sum[get_column_letter(month_num+9)+str(i+2)].fill = overspent



def qqAdd():
    for i in range(2, month_num+1):
        search = wb.get_sheet_by_name(month_list[i])
        add = wb.get_sheet_by_name(month_list[i-1])
        search_max = search.max_row
        add_max = add.max_row
        qqTotal = float(0)
        for j in range(1, search_max+1):
            if search['C'+str(j)].value == 'qq':
                qqTotal += float(search['D'+str(j)].value)
        add['E'+str(add_max+2)] = float(add['E'+str(add_max)].value) + qqTotal

def qqFill():
    for i in range(1, month_num+1):
        blue = wb.get_sheet_by_name(month_list[i])
        for j in range(2, x+1):
            if blue['C'+str(j)].value == 'qq':
                blue['D'+str(j)].fill = PatternFill(start_color='7BBDD3', fill_type='solid')
                blue['E'+str(j)].fill = PatternFill(start_color='7BBDD3', fill_type='solid')


def shortExit():
    if day_num >= 20 and Sum[get_column_letter(month_num+9)+'16'].value == 0:
        if input('Did you already log student loans this month?\n').lower().startswith('n'):
            print('How much?')
            autoAdd('student loans', 'Student Loans', float(input()))
    qqFill()
    overspent_goal()
 #   categoryTotal()
    monthChartBreakdown()
    chartBreakdown()
    print('Calculated %s items' % (items))
    time.sleep(1)
    save_and_update()
    secret()
    return


#Main
year = datetime.datetime.today().strftime('%Y')
#sched = BackgroundScheduler()
month_num = int(datetime.datetime.today().strftime('%m'))
day_num = int(datetime.datetime.today().strftime('%d'))
hour_num = int(datetime.datetime.today().strftime('%H'))
items = 0
first = False

#month_list represents current month for formatting. Jan = 1
month_list = ['Year Total','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sept','Oct','Nov','Dec']
#wb = 'TCM_Final.xlsx'
#year_title = 'TCM_Final'
year_title = year +' Monthly Expenses'
excel = year_title + '.xlsx'
wb = px.load_workbook(excel)
Sum = wb.get_sheet_by_name('Summary')
category_words = wb.get_sheet_by_name('Category Words')
month = setWorkbook()
max_row = month.max_row # x represents the number on the row that you're currently at
month_style = formatSheet(month)
categories = setSummarySheet()
category_dict = setSubCategories()
goals_list = presetGoals()
#print(category_dict)
setMonth()
receipts = False
keyword = ''
if hour_num < 12:
    hour = 'Morning'
elif hour_num >= 12 and hour_num < 17:
    hour = 'Afternoon'
else:
    hour = 'Evening'

#@sched.scheduled_job('cron', day_of_week='mon-fri', hour=12, minute=10)
def routine():
    global receipts
    print('Good %s, Alex. Do you have receipts to add? [Y or N]' % (hour))
    add_receipt = input()
    if add_receipt.lower().startswith('manual'):
        manual_correction(month)
        save_and_update()
    if add_receipt.lower().startswith('goal'):
        changeGoals = goals()
    if add_receipt.lower().startswith('correct'): # or changeGoals.startswith('n'):
        shortExit()
    elif add_receipt.lower().startswith('y'): # or changeGoals.startswith('y'):
        receipts == True
        print()
        reviewMaterials()
        revertStyle()
        orderOfStatements()
    else:
        receipts == False

routine()

#sched.start()
# Shut it down if you make changes to the code
# To stop program from repeating itself, type sched.shutdown(), then restart
